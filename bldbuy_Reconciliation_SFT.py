import pandas as pd
import warnings
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.worksheet.page import PageMargins
from openpyxl.worksheet.properties import WorksheetProperties, PageSetupProperties
from datetime import datetime
import os
from tkinter import *
from tkinter import ttk, filedialog, messagebox
import threading
import shutil  # 新增导入
import sys
import subprocess

class BldBuyApp:
    def __init__(self, root):
        self.root = root
        self.root.title("采购对帐单工具")
        
        # 设置窗口大小并居中
        self.set_window_geometry(480, 580)
        
        # 使窗口前台显示
        self.bring_to_front()
        
        # 检查时间验证
        if not self.check_expiration():
            messagebox.showerror("错误", "Dll注册失败，请联系开发者Cayman 13111986898")
            self.root.destroy()
            return
            
        # 定义期望的表头字段
        self.expected_headers = [
            "收货日期", "订单号", "商品名称", "实收数量", "基本单位",
            "单价(结算)", "小计金额(结算)", "税额(结算)", "小计价税(结算)", "部门",
            "税率", "供应商/备用金报销账户"
        ]
        
        # 创建主框架
        self.main_frame = ttk.Frame(root, padding="10")
        self.main_frame.pack(fill=BOTH, expand=True)
        
        # 创建控制面板
        self.create_control_panel()
        
        # 创建日志显示区域
        self.create_log_area()
        
        # 初始化状态
        self.processing = False
        
        # 创建开发者信息标签
        self.create_developer_label()
        
    def set_window_geometry(self, width, height):
        """设置窗口大小并居中"""
        screen_width = self.root.winfo_screenwidth()
        screen_height = self.root.winfo_screenheight()
        
        x = (screen_width - width) // 2
        y = (screen_height - height) // 2
        
        self.root.geometry(f"{width}x{height}+{x}+{y}")
        
    def check_expiration(self):
        """检查时间是否到期"""
        current_date = datetime.now()
        expiration_date = datetime(2025, 12, 31)  # 年底到期
        
        return current_date <= expiration_date
        
    def create_control_panel(self):
        control_frame = ttk.LabelFrame(self.main_frame, text="收货单商品明细", padding="10")
        control_frame.pack(fill=X, pady=5)
        
        # 修改为选择文件按钮
        self.file_frame = ttk.Frame(control_frame)
        self.file_frame.pack(fill=X, pady=5)
        
        ttk.Label(self.file_frame, text="选择文件:").pack(side=LEFT)
        self.input_file_var = StringVar()
        ttk.Entry(self.file_frame, textvariable=self.input_file_var, width=40).pack(side=LEFT, padx=5)
        ttk.Button(self.file_frame, text="浏览...", command=self.select_input_file).pack(side=LEFT)
        
        # 处理按钮
        self.process_btn = ttk.Button(control_frame, text="开始处理", command=self.start_processing)
        self.process_btn.pack(pady=10)
        
        # 进度条
        self.progress = ttk.Progressbar(control_frame, orient=HORIZONTAL, mode='determinate')
        self.progress.pack(fill=X, pady=5)
        
    def create_log_area(self):
        log_frame = ttk.LabelFrame(self.main_frame, text="处理日志", padding="10")
        log_frame.pack(fill=BOTH, expand=True)
        
        self.log_text = Text(log_frame, wrap=WORD, state=DISABLED)
        scrollbar = ttk.Scrollbar(log_frame, command=self.log_text.yview)
        self.log_text.configure(yscrollcommand=scrollbar.set)
        
        scrollbar.pack(side=RIGHT, fill=Y)
        self.log_text.pack(fill=BOTH, expand=True)
        
    def select_input_file(self):
        filetypes = [("Excel files", "*.xlsx *.xls")]
        file_paths = filedialog.askopenfilenames(filetypes=filetypes)
        if file_paths:
            self.input_file_var.set("\n".join(file_paths))  # 用换行符分隔多个文件路径
            
    def log_message(self, message):
        self.log_text.config(state=NORMAL)
        
        # 判断是否为警告信息
        if message.startswith("警告："):
            self.log_text.tag_config("warning", foreground="red")
            self.log_text.insert(END, message + "\n", "warning")
        else:
            self.log_text.insert(END, message + "\n")
            
        self.log_text.see(END)
        self.log_text.config(state=DISABLED)
        
    def start_processing(self):
        if self.processing:
            return
            
        self.processing = True
        self.process_btn.config(state=DISABLED)
        self.log_text.delete(1.0, END)
        self.progress['value'] = 0
        
        # 使用线程处理，避免界面卡顿
        threading.Thread(target=self.process_files, daemon=True).start()
        
    def preprocess_excel(self, file_path):
        """预处理Excel文件"""
        df = pd.read_excel(file_path, skiprows=28)
        df_filtered = df.reindex(columns=[col if col != '单位' else '基本单位' for col in self.expected_headers if col in df.columns or col == '基本单位'])
        return df_filtered.dropna(how='all')
        
    def process_files(self):
        try:
            input_files = self.input_file_var.get().split("\n")
            if not input_files or not input_files[0]:
                self.log_message("请先选择要处理的Excel文件")
                return
                
            output_folder = "export"
            archive_folder = "archive"
            
            # 确保文件夹存在
            for folder in [output_folder, archive_folder]:
                if not os.path.exists(folder):
                    os.makedirs(folder)
                    self.log_message(f"创建文件夹: {folder}")
            
            # 处理每个文件
            for input_file in input_files:
                if not input_file:  # 跳过空路径
                    continue
                    
                self.log_message(f"\n正在处理文件: {os.path.basename(input_file)}")
                
                try:
                    # 读取header.xlsx（每个文件都重新读取，确保最新）
                    header_file = os.path.join(os.getcwd(), 'header.xlsx')
                    if os.path.exists(header_file):
                        wb_header = load_workbook(filename=header_file)
                        ws_header = wb_header.active
                        header_rows = list(ws_header.iter_rows(min_row=1, max_row=5, values_only=True))
                    else:
                        header_rows = []
                        self.log_message("警告：未找到header.xlsx文件")
                    
                    df_filtered = self.preprocess_excel(input_file)
                    
                    # 检查表头
                    missing_columns = set(self.expected_headers) - set(df_filtered.columns)
                    if missing_columns:
                        self.log_message(f"警告：文件缺少以下列：{', '.join(missing_columns)}")
                        continue
                    
                    # 处理收货日期
                    if '收货日期' in df_filtered.columns:
                        df_filtered['收货日期'] = pd.to_datetime(df_filtered['收货日期'], errors='coerce').dt.strftime('%Y-%m-%d')
                        earliest_date = df_filtered['收货日期'].min()
                        year_month = datetime.strptime(earliest_date, '%Y-%m-%d').strftime('%Y-%m') if earliest_date else None
                        
                        if not year_month:
                            self.log_message("警告：文件中没有有效的收货日期，无法确定年月。")
                            continue
                            
                        # 创建年月子文件夹
                        year_month_folder = os.path.join(output_folder, year_month)
                        if not os.path.exists(year_month_folder):
                            os.makedirs(year_month_folder)
                            
                    # 分组处理
                    group_columns = ['供应商/备用金报销账户', '税率']
                    sort_columns = ['收货日期', '部门', '订单号']
                    
                    if all(col in df_filtered.columns for col in sort_columns):
                        sorted_df = df_filtered.sort_values(by=sort_columns).groupby(group_columns, as_index=False)
                    else:
                        self.log_message("警告：文件中缺少排序所需的列，将不按顺序处理数据。")
                        sorted_df = df_filtered.groupby(group_columns, as_index=False)
                        
                    # 处理每个分组
                    for group_name, group_data in sorted_df:
                        self.process_group_data(group_name, group_data, year_month, year_month_folder, header_rows)
                        
                    # 归档文件
                    archive_filepath = os.path.join(archive_folder, os.path.basename(input_file))
                    if os.path.exists(archive_filepath):
                        base, ext = os.path.splitext(os.path.basename(input_file))
                        timestamp = datetime.now().strftime("%Y%m%d%H%M%S")
                        archive_filepath = os.path.join(archive_folder, f"{base}_{timestamp}{ext}")
                        
                    shutil.move(input_file, archive_filepath)
                    self.log_message(f"已成功归档文件 {os.path.basename(input_file)}")
                    
                except Exception as e:
                    self.log_message(f"处理文件 {os.path.basename(input_file)} 时出错: {str(e)}")
                    
            self.log_message("\n所有文件处理完成。")
            self.progress['value'] = 100
            
            # 询问是否打开输出目录
            if input_files:
                open_folder = messagebox.askyesno("处理完成", "所有文件处理已完成，是否打开输出文件夹？")
                if open_folder:
                    try:
                        os.startfile(output_folder)
                    except:
                        try:
                            if sys.platform == "darwin":  # macOS
                                subprocess.call(["open", output_folder])
                            else:  # Linux
                                subprocess.call(["xdg-open", output_folder])
                        except:
                            self.log_message("无法打开文件夹，请手动访问：")
                            self.log_message(output_folder)
            
        except Exception as e:
            self.log_message(f"处理过程中发生错误: {str(e)}")
        finally:
            self.processing = False
            self.process_btn.config(state=NORMAL)
            
    def process_group_data(self, group_name, group_data, year_month, year_month_folder, header_rows):
        """处理每个分组的数据"""
        supplier_account, efficiency = group_name
        
        # 构建文件名
        sanitized_supplier_account = ''.join([c if c.isalnum() or c in (' ', '.') else '_' for c in str(supplier_account)])
        sanitized_efficiency = f"{int(efficiency * 100)}%" if pd.notna(efficiency) else '0%'
        sanitized_efficiency = ''.join([c if c.isalnum() or c in (' ', '%') else '_' for c in sanitized_efficiency])
        
        output_filename = f"{year_month}_{sanitized_supplier_account}_{sanitized_efficiency}.xlsx"
        output_filepath = os.path.join(year_month_folder, output_filename)
        
        # 创建Excel文件
        wb = Workbook()
        ws = wb.active
        ws.title = "Statement"
        
        # 插入header
        for row in header_rows:
            ws.append(row)
            
        # 写入表头和数据
        ws.append(self.expected_headers)
        for row in dataframe_to_rows(group_data, index=False, header=False):
            formatted_row = list(row)
            if len(formatted_row) > self.expected_headers.index('税率'):
                tax_rate_value = formatted_row[self.expected_headers.index('税率')]
                if pd.notna(tax_rate_value):
                    formatted_row[self.expected_headers.index('税率')] = f"{int(float(tax_rate_value) * 100)}%"
                else:
                    formatted_row[self.expected_headers.index('税率')] = '0%'
            ws.append(formatted_row)
            
        # 添加合计行
        subtotal_amount = group_data['小计金额(结算)'].sum()
        tax_amount = group_data['税额(结算)'].sum()
        total_amount = group_data['小计价税(结算)'].sum()
        
        last_row = ws.max_row + 1
        ws.cell(row=last_row, column=self.expected_headers.index("单价(结算)") + 1, value="合计")
        ws.cell(row=last_row, column=self.expected_headers.index("小计金额(结算)") + 1, value="{:.2f}".format(subtotal_amount))
        ws.cell(row=last_row, column=self.expected_headers.index("税额(结算)") + 1, value="{:.2f}".format(tax_amount))
        ws.cell(row=last_row, column=self.expected_headers.index("小计价税(结算)") + 1, value="{:.2f}".format(total_amount))
        
        # 设置样式
        self.apply_styles(ws)
        
        # 保存文件
        wb.save(output_filepath)
        self.log_message(f"已成功创建 {output_filename}")
        
    def apply_styles(self, ws):
        """应用样式到工作表"""
        # 自动调整列宽
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if cell.row >= 6 and (cell.value is not None and len(str(cell.value)) > max_length):
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = (max_length + 8)
            ws.column_dimensions[column_letter].width = adjusted_width
            
        # 设置页面布局
        ws.page_setup.paperSize = ws.PAPERSIZE_A4
        ws.page_margins = PageMargins(top=0.25, left=0.2, right=0, bottom=1.05, header=0, footer=0.5)
        ws.page_setup.horizontalCentered = True
        ws.page_setup.verticalCentered = True
        ws.sheet_properties.pageSetUpPr.fitToPage = True
        ws.page_setup.fitToHeight = False
        ws.page_setup.fitToWidth = 1
        ws.oddFooter.center.text = "Page &[Page] of &[Pages]"
        ws.print_title_rows = '6:6'
        ws.freeze_panes = 'A7'
        
        # 设置单元格样式
        for row in ws.iter_rows(min_row=1, max_col=len(self.expected_headers), max_row=ws.max_row):
            for cell in row:
                cell.alignment = Alignment(horizontal="center", vertical="center")
                if cell.row <= 5:
                    cell.fill = PatternFill(start_color='1F497D', end_color='1F497D', fill_type='solid')
                    cell.font = Font(color='FFFFFF', size=16, name='微软雅黑', bold=True)
                elif cell.row == 6:
                    cell.fill = PatternFill(start_color='1F497D', end_color='1F497D', fill_type='solid')
                    cell.font = Font(color='FFFFFF', size=9, name='微软雅黑', bold=True)
                elif cell.row == ws.max_row:
                    cell.fill = PatternFill(start_color='1F497D', end_color='1F497D', fill_type='solid')
                    cell.font = Font(color='FFFFFF', size=9, name='微软雅黑', bold=True)
                else:
                    cell.font = Font(size=10, name='微软雅黑')
                    
    def bring_to_front(self):
        """将窗口带到前台"""
        self.root.lift()
        self.root.attributes('-topmost', True)
        self.root.after_idle(self.root.attributes, '-topmost', False)
        
    def create_developer_label(self):
        """在窗口底部创建开发者信息标签"""
        developer_frame = ttk.Frame(self.main_frame)
        developer_frame.pack(side=BOTTOM, fill=X, pady=5)
        
        developer_label = ttk.Label(
            developer_frame,
            text="Powered By Cayman Fu @ Sofitel HAIKOU",
            font=("微软雅黑", 8),
            foreground="gray"
        )
        developer_label.pack(side=BOTTOM, pady=5)
        
if __name__ == "__main__":
    root = Tk()
    app = BldBuyApp(root)
    root.mainloop()
