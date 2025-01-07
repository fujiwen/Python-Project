import pandas as pd
import os
import re
from datetime import datetime, timedelta
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, NamedStyle
import sys
import time

# 获取当前脚本所在的目录
def resource_path(relative_path):
    """ Get absolute path to resource, works for dev and for PyInstaller """
    try:
        # PyInstaller creates a temp folder and stores path in _MEIPASS
        base_path = sys._MEIPASS
    except Exception:
        base_path = os.path.abspath(".")

    return os.path.join(base_path, relative_path)

# 构建import目录路径
current_dir = os.path.dirname(os.path.abspath(sys.executable))
import_dir = os.path.join(current_dir, 'import')

# 检查import目录是否存在
if not os.path.exists(import_dir):
    print(f"Error: Directory {import_dir} does not exist.")
    sys.exit(1)

# 初始化一个空的DataFrame用于存储所有整理后的数据
all_data = []

# 定义需要处理的工作表名称（如果已知）
sheets_to_process = ['Aged Reports']  # 根据实际工作表名称调整

# 遍历import目录下的所有.xlsm文件
for file_name in os.listdir(import_dir):
    if file_name.endswith('.xlsm'):
        input_file_path = os.path.join(import_dir, file_name)
        
        # 读取Excel文件中的所有工作表
        xls = pd.ExcelFile(input_file_path)

        for sheet_name in sheets_to_process:
            # 读取每个工作表的数据，跳过前两行
            df = pd.read_excel(xls, sheet_name=sheet_name, skiprows=2)
            
            # 确保数值列的类型是float
            numeric_columns = ['Total', '30 days', '60 days', '90 days', '120 days', '150 days', '180 days']
            df[numeric_columns] = df[numeric_columns].apply(pd.to_numeric, errors='coerce')
            
            
            # 查找并删除包含'Total'的行，这里使用'Transaction Date'和'Transaction Reference'列来查找
            if 'Transaction Date' in df.columns and 'Transaction Reference' in df.columns:
                # 使用逻辑或操作符 (|) 来检查任一列是否包含'Total'
                df = df[~df['Transaction Date'].astype(str).str.lower().str.contains('total') &
                        ~df['Transaction Reference'].astype(str).str.lower().str.contains('total')]

            # 检查'Transaction Date'列，并处理非日期格式的内容
            if 'Transaction Date' in df.columns:
                supplier_id_col = 'Supplier ID'
                
                # 如果没有'Supplier ID'列，创建一个新的
                if supplier_id_col not in df.columns:
                    df[supplier_id_col] = None
                
                # 遍历'Transaction Date'列，查找非日期格式的内容并移到'Supplier ID'列
                for index, row in df.iterrows():
                    try:
                        date = pd.to_datetime(row['Transaction Date']).date()  # 只保留年月日
                        df.at[index, 'Transaction Date'] = date
                    except ValueError:
                        df.at[index, supplier_id_col] = row['Transaction Date']
                        df.at[index, 'Transaction Date'] = pd.NaT  # 将原始位置设置为缺失值

            # 检查'Transaction Reference'列，并处理不符合条件的内容
            if 'Transaction Reference' in df.columns:
                supplier_name_col = 'Supplier Name'

                # 如果没有'Supplier Name'列，创建一个新的
                if supplier_name_col not in df.columns:
                    df[supplier_name_col] = None

                # 遍历'Transaction Reference'列，查找不符合条件的内容并移到'Supplier Name'列
                for index, row in df.iterrows():
                    if not re.match(r'^[A-Za-z0-9\-]*$', str(row['Transaction Reference'])):
                        df.at[index, supplier_name_col] = row['Transaction Reference']
                        df.at[index, 'Transaction Reference'] = None  # 清除原始位置内容

            # 将处理过的数据添加到all_data列表中
            all_data.append(df)

        # 关闭Excel文件句柄
        xls.close()

# 合并所有整理后的数据
final_df = pd.concat(all_data, ignore_index=True)

# 自动向下填充空白格，直到遇到新的Supplier ID和Supplier Name
supplier_cols = ['Supplier ID', 'Supplier Name']

# 使用 .ffill() 方法来代替 .fillna(method='ffill')
final_df[supplier_cols] = final_df[supplier_cols].ffill()

# 构建输出文件路径
output_file_path = os.path.join(current_dir, 'cleaned_data.xlsx')

# 将数据写入新的Excel文件
with pd.ExcelWriter(output_file_path, engine='openpyxl') as writer:
    final_df.to_excel(writer, sheet_name='Cleaned Aged Reports', index=False)

print(f"Data has been cleaned and saved to {output_file_path}")

# 读取Excel文件
file_path = output_file_path
df = pd.read_excel(file_path)

# 确保'Transaction Date'是datetime类型
df['Transaction Date'] = pd.to_datetime(df['Transaction Date'], errors='coerce')

# 按月和'Supplier ID', 'Supplier Name'分组，并合计'Total'
df['YearMonth'] = df['Transaction Date'].dt.to_period('M')
grouped = df.groupby(['Supplier ID', 'Supplier Name', 'YearMonth']).agg(
    Total_Transactions=('Total', 'sum')
).reset_index()

# 将Period类型的'YearMonth'转换为字符串格式
grouped['YearMonth'] = grouped['YearMonth'].astype(str)

# 构造透视表，以'Supplier ID', 'Supplier Name'为索引，以'YearMonth'为列，合计'Total_Transactions'
pivot_table = grouped.pivot_table(index=['Supplier ID', 'Supplier Name'],
                                  columns='YearMonth',
                                  values='Total_Transactions',
                                  aggfunc='sum').fillna(0)

# 对列名（即年月）按照从近到远排序
sorted_columns = sorted(pivot_table.columns, key=lambda x: datetime.strptime(x, '%Y-%m'), reverse=True)
sorted_pivot_table = pivot_table[sorted_columns]

# 在这里直接操作透视表的副本
sorted_pivot_table = sorted_pivot_table.copy()
sorted_pivot_table['Total_Sum'] = sorted_pivot_table.sum(axis=1)

# 将透视表重置索引以便'Supplier ID', 'Supplier Name'成为普通列
result_df = sorted_pivot_table.reset_index()

# 将表格的0转换为"-"，但保留"总合计"列中的数值
result_df = result_df.apply(lambda x: x.replace({0: '-'}) if x.name != 'Total_Sum' else x)

# 移动'Total_Sum'列到年月数据的最前面
total_sum_col = result_df.pop('Total_Sum')
year_month_cols = [col for col in result_df.columns if col not in ['Supplier ID', 'Supplier Name']]
result_df.insert(result_df.columns.get_loc('Supplier Name') + 1, 'Total_Sum', total_sum_col)

# 找到最新的年月
latest_yearmonth = sorted_columns[0]

# 构建输出文件路径，并确保文件名唯一
styled_output_base_path = os.path.join(current_dir, f"{latest_yearmonth}_AP Aging Report")
styled_output_file_extension = ".xlsx"
styled_output_file_path = styled_output_base_path + styled_output_file_extension

# 检查文件是否存在并重命名
counter = 1
while os.path.exists(styled_output_file_path):
    styled_output_file_path = f"{styled_output_base_path}_{counter}{styled_output_file_extension}"
    counter += 1

# 写入新的Excel文件
with pd.ExcelWriter(styled_output_file_path, engine='openpyxl') as writer:
    # 添加一个空行到结果 DataFrame 的末尾
    empty_row = pd.DataFrame(columns=result_df.columns)
    result_df_with_empty_row = pd.concat([result_df, empty_row], ignore_index=True)
    
    result_df_with_empty_row.to_excel(writer, index=False, sheet_name='Aggregated Data')

    # 加载工作簿和工作表
    workbook = writer.book
    worksheet = writer.sheets['Aggregated Data']

    # 设置表头样式
    header_fill = PatternFill(start_color="00009B", end_color="00009B", fill_type="solid")
    header_font_bold = Font(name='微软雅黑', size=9, color='FFFFFF', bold=True)  # 表头字体加粗
    for cell in worksheet[1]:
        cell.fill = header_fill
        cell.font = header_font_bold

    # 定义会计专用样式
    accounting_style = NamedStyle(name="accounting", number_format='#,##0.00;[Red]-#,##0.00')
    accounting_style.font = Font(name='微软雅黑', size=10)
    accounting_style.alignment = Alignment(horizontal='right')  # 设置右对齐

    # 插入天数信息行
    days_above_headers = []
    for idx, col in enumerate(worksheet.iter_cols(min_row=1, max_row=1, min_col=4, max_col=len(sorted_columns)+2, values_only=True), start=1):
        days_above_headers.append(idx * 30)  # 30, 60, 90...

    # 插入一行用于天数信息，并设置自定义格式
    custom_format = NamedStyle(name="custom_days", number_format='0 "Days"')
    custom_format.font = Font(name='微软雅黑', size=10)
    custom_format.alignment = Alignment(horizontal='center')

    # 插入新行并在其中填写天数信息，从第四列开始
    worksheet.insert_rows(1)
    for idx, days in enumerate(days_above_headers, start=4):  # 从第四列开始（跳过前三列）
        cell = worksheet.cell(row=1, column=idx, value=days)
        cell.style = custom_format

    # 插入空白行（表头下方），并计算每列的数据合计
    data_start_row = 3  # 数据开始的行号（考虑了天数信息行）
    data_end_row = worksheet.max_row  # 数据结束的行号
    col_start_idx = 3  # 合计开始的列索引（第三列）

    # 插入空白行
    worksheet.insert_rows(data_start_row)

    # 定义合计行样式
    total_row_style = NamedStyle(name="total_row_style")
    total_row_style.fill = PatternFill(start_color="DDEBF7", end_color="DDEBF7", fill_type="solid")
    total_row_style.font = Font(name='微软雅黑', size=11, color='002060')
    total_row_style.alignment = Alignment(vertical='center', horizontal='right')
    total_row_style.number_format = '#,##0.00;[Red]-#,##0.00'  # 会计专用格式，保留两位小数

    # 计算合计并填入空白行，并应用样式
    for col in worksheet.iter_cols(min_row=data_start_row+1, max_row=data_end_row+1, min_col=col_start_idx, max_col=worksheet.max_column):
        sum_value = sum(cell.value for cell in col if isinstance(cell.value, (int, float)))
        sum_cell = worksheet.cell(row=data_start_row, column=col[0].column, value=sum_value)
        sum_cell.style = total_row_style  # 应用合计行样式

    # 设置主体数据样式
    for row in worksheet.iter_rows(min_row=data_start_row+1, max_row=worksheet.max_row, min_col=1, max_col=worksheet.max_column):
        for cell in row:
            cell.font = Font(name='微软雅黑', size=10)
            cell.alignment = Alignment(horizontal='right')  # 主体数据右对齐
            if isinstance(cell.value, (int, float)):  # 如果是数值，则应用会计专用样式
                cell.style = accounting_style
            elif isinstance(cell.value, str) and cell.value == '0':  # 如果值是字符'0'，则替换为'-'
                cell.value = '-'

    # 自定义列宽设置
    worksheet.column_dimensions['A'].width = 15  # 第一列
    worksheet.column_dimensions['B'].width = 40  # 第二列
    
    # 设置其他列宽为15
    for idx, col in enumerate(worksheet.columns, start=1):
        if idx > 2:  # 从第三列开始
            worksheet.column_dimensions[col[0].column_letter].width = 20

    # 设置所有行的高度为22.5磅
    for row in worksheet.iter_rows(min_row=1, max_row=worksheet.max_row):
        worksheet.row_dimensions[row[0].row].height = 22.5

    # 冻结表格前两行
    worksheet.freeze_panes = worksheet['A4']  # 冻结前两行

    # 取消表格网格线
    worksheet.sheet_view.showGridLines = False

print(f"Styled aggregated data has been written to {styled_output_file_path} with customized column widths, all row heights set to 22.5pt, grid lines removed, the first two rows frozen, a blank row inserted below the header with column totals from the third column onwards, and styled according to specifications including accounting format with two decimal places.")

# 删除cleaned_data.xlsx文件
if os.path.exists(output_file_path):
    os.remove(output_file_path)
    print(f"Deleted {output_file_path}")

# 等待5秒
print("Waiting for 5 seconds before deleting files in the import directory...")
time.sleep(5)

# 删除import目录内的所有文件
try:
    for filename in os.listdir(import_dir):
        file_path = os.path.join(import_dir, filename)
        if os.path.isfile(file_path):
            try:
                os.remove(file_path)
                print(f"Deleted file: {file_path}")
            except PermissionError as e:
                print(f"Permission denied to delete {file_path}: {e}")
            except Exception as e:
                print(f"Failed to delete {file_path}: {e}")
except Exception as e:
    print(f"Failed to delete files in the import directory: {e}")

