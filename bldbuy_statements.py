import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.worksheet.page import PageMargins
from openpyxl.worksheet.properties import WorksheetProperties, PageSetupProperties
from datetime import datetime
import os

# 定义期望的表头字段及其顺序，并直接将"单位"更改为"基本单位"
expected_headers = [
    "收货日期", "订单号", "商品名称", "实收数量", "基本单位", 
    "单价(结算)", "小计金额(结算)", "税额(结算)", "小计价税(结算)", "部门",
    "税率", "供应商/备用金报销账户"  # 互换位置
]

# 确保 'import' 文件夹存在
input_folder = "import"
if not os.path.exists(input_folder):
    os.makedirs(input_folder)

# 确保 'export' 文件夹存在
output_folder = "export"
if not os.path.exists(output_folder):
    os.makedirs(output_folder)

# 读取header.xlsx中的前五行
header_file = os.path.join(os.getcwd(), 'header.xlsx')
if os.path.exists(header_file):
    wb_header = load_workbook(filename=header_file)
    ws_header = wb_header.active
    header_rows = list(ws_header.iter_rows(min_row=1, max_row=5, values_only=True))
else:
    header_rows = []

# 读取上传的Excel文件（假设所有Excel文件都在import文件夹中）
input_files = [f for f in os.listdir(input_folder) if f.endswith('.xlsx') or f.endswith('.xls')]
for input_file in input_files:
    df = pd.read_excel(os.path.join(input_folder, input_file))

    # 确保所有预期的表头都存在于输入文件中，并按照指定的顺序重排列
    df_filtered = df.reindex(columns=[col if col != '单位' else '基本单位' for col in expected_headers if col in df.columns or col == '基本单位'])

    # 格式化收货日期为 YYYY-MM-DD
    df_filtered['收货日期'] = pd.to_datetime(df_filtered['收货日期']).dt.strftime('%Y-%m-%d')

    # 按指定条件分组并排序
    group_columns = ['供应商/备用金报销账户', '税率']  # 假设'税率'代表效率
    sort_columns = ['订单号']

    # 先排序后分组
    sorted_df = df_filtered.sort_values(by=sort_columns).groupby(group_columns, as_index=False)

    # 遍历每个分组并保存为单独的文件
    for group_name, group_data in sorted_df:
        supplier_account, efficiency = group_name  # 解构分组键
        
        # 构建文件名，去除非法字符并替换为空格
        sanitized_supplier_account = ''.join([c if c.isalnum() or c in (' ', '.') else '_' for c in str(supplier_account)])
        
        # 将税率转换为整数百分比格式用于文件名
        sanitized_efficiency = f"{int(efficiency * 100)}%" if pd.notna(efficiency) else '0%'
        sanitized_efficiency = ''.join([c if c.isalnum() or c in (' ', '%') else '_' for c in sanitized_efficiency])
        
        # 构造文件路径和名称
        output_filename = f"{sanitized_supplier_account}_{sanitized_efficiency}.xlsx"
        output_filepath = os.path.join(output_folder, output_filename)
        
        # 创建一个新的工作簿对象
        wb = Workbook()
        ws = wb.active
        ws.title = "Statement"

        # 插入header.xlsx中的前五行
        for row in header_rows:
            ws.append(row)

        # 写入表头信息
        ws.append(expected_headers)

        # 写入分组的数据
        for row in dataframe_to_rows(group_data, index=False, header=False):
            formatted_row = list(row)
            # 将税率转换为整数百分比格式，即使是0也要显示为0%
            if len(formatted_row) > expected_headers.index('税率'):
                tax_rate_value = formatted_row[expected_headers.index('税率')]
                if pd.notna(tax_rate_value):  # 确保不是NaN
                    formatted_row[expected_headers.index('税率')] = f"{int(float(tax_rate_value) * 100)}%" if tax_rate_value is not None else '0%'
                else:
                    formatted_row[expected_headers.index('税率')] = '0%'
            ws.append(formatted_row)

        # 计算合计
        subtotal_amount = group_data['小计金额(结算)'].sum()
        tax_amount = group_data['税额(结算)'].sum()
        total_amount = group_data['小计价税(结算)'].sum()

        # 添加合计行到最后一行
        last_row = ws.max_row + 1
        ws.cell(row=last_row, column=expected_headers.index("单价(结算)") + 1, value="合计")
        ws.cell(row=last_row, column=expected_headers.index("小计金额(结算)") + 1, value="{:.2f}".format(subtotal_amount))
        ws.cell(row=last_row, column=expected_headers.index("税额(结算)") + 1, value="{:.2f}".format(tax_amount))
        ws.cell(row=last_row, column=expected_headers.index("小计价税(结算)") + 1, value="{:.2f}".format(total_amount))

        # 自动调整列宽
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter  # 获取列字母
            for cell in column:
                try:
                    if cell.row >= 6 and (cell.value is not None and len(str(cell.value)) > max_length):  # 忽略前五行，从第六行开始计算
                        max_length = len(str(cell.value))
                except TypeError:
                    pass  # 忽略无法计算长度的单元格值（如None）
            adjusted_width = (max_length + 8)  # 添加一些额外空间
            ws.column_dimensions[column_letter].width = adjusted_width

        # 设置页边距和水平居中
        ws.page_margins = PageMargins(top=1.0, left=0.2, right=0.2, bottom=1.8)
        ws.page_setup.horizontalCentered = True
        ws.page_setup.verticalCentered = True

        # 设置打印缩放，使所有列打印在一页上
        ws.sheet_properties.pageSetUpPr.fitToPage = True
        ws.page_setup.fitToHeight = False  # 不适应高度，只适应宽度
        ws.page_setup.fitToWidth = 1  # 所有列打印在一页上

        # 添加页脚（页码）
        ws.oddFooter.center.text = "Page &[Page] of &[Pages]"

       # 设置打印标题行，使得表头在每一页都打印出来
        ws.print_title_rows = '6:6'  # 表头在第六行

        # 格式化单元格样式
        for row in ws.iter_rows(min_row=1, max_col=len(expected_headers), max_row=ws.max_row):
            for cell in row:
                cell.alignment = Alignment(horizontal="center", vertical="center")
                if cell.row <= 5:  # Header行格式
                    cell.fill = PatternFill(start_color='1F497D', end_color='1F497D', fill_type='solid')  # 轻灰色填充
                    cell.font = Font(color='FFFFFF', size=16, name='微软雅黑', bold=True)
                elif cell.row == 6:  # 表头格式
                    cell.fill = PatternFill(start_color='1F497D', end_color='1F497D', fill_type='solid')
                    cell.font = Font(color='FFFFFF', size=9, name='微软雅黑', bold=True)
                elif cell.row == ws.max_row:  # 合计行格式
                    cell.fill = PatternFill(start_color='1F497D', end_color='1F497D', fill_type='solid')
                    cell.font = Font(color='FFFFFF', size=9, name='微软雅黑', bold=True)
                else:  # 分组数据格式
                    cell.font = Font(size=10, name='微软雅黑')

        # 保存文件
        wb.save(output_filepath)

        print(f"已成功创建 {output_filepath}")
        
# 确保 'archive' 文件夹存在
archive_folder = os.path.join(os.getcwd(), "archive")
if not os.path.exists(archive_folder):
    os.makedirs(archive_folder)

# 将处理过的文件从 'import' 文件夹移动到 'archive' 文件夹
for input_file in input_files:
    original_filepath = os.path.join(input_folder, input_file)
    archive_filepath = os.path.join(archive_folder, input_file)

    # 检查文件是否存在
    if os.path.exists(original_filepath):
        # 为了避免覆盖同名文件，检查目标路径是否存在，如果存在则重命名
        if os.path.exists(archive_filepath):
            base, ext = os.path.splitext(input_file)
            timestamp = datetime.now().strftime("%Y%m%d%H%M%S")
            archive_filepath = os.path.join(archive_folder, f"{base}_{timestamp}{ext}")

        try:
            os.rename(original_filepath, archive_filepath)
            print(f"已成功归档文件 {input_file} 至 {archive_folder}")
        except Exception as e:
            print(f"归档文件 {input_file} 失败: {e}")
    else:
        print(f"文件 {original_filepath} 不存在，无法归档。")

print("所有文件处理完成。")
print("所有文件处理完成。")
