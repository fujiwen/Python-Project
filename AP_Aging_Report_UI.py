import tkinter as tk
from tkinter import ttk, filedialog, messagebox, Text
import threading
import pandas as pd
import os
import re
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, NamedStyle
import time

class AP_Aging_Report_App:
    def __init__(self, root, master_window):
        self.root = root
        self.master = master_window  # 使用传入的主窗口对象
        self.master.title("AP 帐龄报表工具")  # 在顶层窗口设置标题
        self.master.geometry("800x600")
        
        # 创建主框架
        self.main_frame = ttk.Frame(root, padding="20")
        self.main_frame.pack(fill=tk.BOTH, expand=True)
        
        # 文件选择和进度条框架
        self.file_frame = ttk.LabelFrame(self.main_frame, text="选择要处理的帐龄报表")
        self.file_frame.pack(fill=tk.X, pady=5)
        
        # 文件选择部分    
        # 添加文字提示
        self.file_label = ttk.Label(self.file_frame, text="选择文件：")
        self.file_label.grid(row=0, column=0, sticky='w', padx=(5, 0), pady=5)
        
        self.file_entry = ttk.Entry(self.file_frame)
        self.file_entry.grid(row=0, column=1, sticky='ew', padx=(5, 0), pady=5)
        
        self.browse_btn = ttk.Button(self.file_frame, text="浏览...", command=self.select_file, width=10)
        self.browse_btn.grid(row=0, column=2, padx=(0, 5), pady=5)
        
        # 开始处理按钮（移动到文件选择框架内）
        self.process_btn = ttk.Button(self.file_frame, text="开始处理", command=self.start_processing)
        self.process_btn.grid(row=1, column=0, columnspan=2, sticky='w', pady=(5, 10), padx=5)
        
        # 配置列权重
        self.file_frame.columnconfigure(0, weight=0)  # 标签列不扩展
        self.file_frame.columnconfigure(1, weight=1)  # 输入框扩展
        self.file_frame.columnconfigure(2, weight=0)  # 按钮列不扩展
        
        # 日志显示
        self.log_frame = ttk.LabelFrame(self.main_frame, text="处理日志")
        self.log_frame.pack(fill=tk.BOTH, expand=True)
        
        # 创建滚动条
        self.log_scroll = ttk.Scrollbar(self.log_frame)
        self.log_scroll.pack(side=tk.RIGHT, fill=tk.Y)
        
        # 在log_frame内添加Text部件用于显示日志
        if not hasattr(self, 'log_text'):
            self.log_text = Text(self.log_frame, height=5, width=50, yscrollcommand=self.log_scroll.set)
            self.log_text.pack(fill='both', expand=True)
        
        # 配置滚动条
        self.log_scroll.config(command=self.log_text.yview)
        
        # 初始化变量
        self.input_file = ""
        self.processing = False
        
    def select_file(self):
        filetypes = [("Excel files", "*.xlsm *.xlsx")]
        file = filedialog.askopenfilename(filetypes=filetypes)
        if file:
            self.file_entry.delete(0, tk.END)
            self.file_entry.insert(0, file)
            self.input_file = file
            
    def log_message(self, message):
        self.log_text.insert(tk.END, message + "\n")
        self.log_text.see(tk.END)
        self.root.update_idletasks()
        
    def start_processing(self):
        if not self.input_file:
            messagebox.showwarning("警告", "请先选择要处理的Excel文件")
            return
            
        if self.processing:
            return
            
        self.processing = True
        self.process_btn.config(state=tk.DISABLED)
        self.log_message("开始处理...")
        
        # 在新线程中运行处理逻辑
        processing_thread = threading.Thread(target=self.run_processing)
        processing_thread.start()
        
    def run_processing(self):
        try:
            # 获取程序所在目录
            current_dir = os.path.dirname(os.path.abspath(__file__))
            
            # 初始化数据
            all_data = []
            sheets_to_process = ['Aged Reports']
            
            # 读取Excel文件
            self.log_message("正在读取文件...")
            xls = pd.ExcelFile(self.input_file)
            
            for sheet_name in sheets_to_process:
                # 读取工作表数据
                df = pd.read_excel(xls, sheet_name=sheet_name, skiprows=2)
                
                # 处理数值列
                numeric_columns = ['Total', '30 days', '60 days', '90 days', '120 days', '150 days', '180 days']
                df[numeric_columns] = df[numeric_columns].apply(pd.to_numeric, errors='coerce')
                
                # 处理Total行
                if 'Transaction Date' in df.columns and 'Transaction Reference' in df.columns:
                    df = df[~df['Transaction Date'].astype(str).str.lower().str.contains('total') &
                            ~df['Transaction Reference'].astype(str).str.lower().str.contains('total')]
                
                # 处理Transaction Date列
                if 'Transaction Date' in df.columns:
                    supplier_id_col = 'Supplier ID'
                    if supplier_id_col not in df.columns:
                        df[supplier_id_col] = None
                    
                    for index, row in df.iterrows():
                        try:
                            date = pd.to_datetime(row['Transaction Date']).date()
                            df.at[index, 'Transaction Date'] = date
                        except ValueError:
                            df.at[index, supplier_id_col] = row['Transaction Date']
                            df.at[index, 'Transaction Date'] = pd.NaT
                
                # 处理Transaction Reference列
                if 'Transaction Reference' in df.columns:
                    supplier_name_col = 'Supplier Name'
                    if supplier_name_col not in df.columns:
                        df[supplier_name_col] = None
                    
                    for index, row in df.iterrows():
                        if not re.match(r'^[A-Za-z0-9\-]*$', str(row['Transaction Reference'])):
                            df.at[index, supplier_name_col] = row['Transaction Reference']
                            df.at[index, 'Transaction Reference'] = None
                
                all_data.append(df)
            
            xls.close()
            
            # 合并数据
            self.log_message("正在合并数据...")
            final_df = pd.concat(all_data, ignore_index=True)
            
            # 填充空白格
            supplier_cols = ['Supplier ID', 'Supplier Name']
            final_df[supplier_cols] = final_df[supplier_cols].ffill()
            
            # 生成透视表
            self.log_message("正在生成透视表...")
            final_df['Transaction Date'] = pd.to_datetime(final_df['Transaction Date'], errors='coerce')
            final_df['YearMonth'] = final_df['Transaction Date'].dt.to_period('M')
            
            grouped = final_df.groupby(['Supplier ID', 'Supplier Name', 'YearMonth']).agg(
                Total_Transactions=('Total', 'sum')
            ).reset_index()
            
            grouped['YearMonth'] = grouped['YearMonth'].astype(str)
            
            pivot_table = grouped.pivot_table(index=['Supplier ID', 'Supplier Name'],
                                            columns='YearMonth',
                                            values='Total_Transactions',
                                            aggfunc='sum').fillna(0)
            
            # 排序并计算合计
            sorted_columns = sorted(pivot_table.columns, key=lambda x: datetime.strptime(x, '%Y-%m'), reverse=True)
            sorted_pivot_table = pivot_table[sorted_columns]
            sorted_pivot_table = sorted_pivot_table.copy()
            sorted_pivot_table['Total_Sum'] = sorted_pivot_table.sum(axis=1)
            
            # 生成最终结果
            result_df = sorted_pivot_table.reset_index()
            result_df = result_df.apply(lambda x: x.replace({0: '-'}) if x.name != 'Total_Sum' else x)
            
            # 移动Total_Sum列
            total_sum_col = result_df.pop('Total_Sum')
            year_month_cols = [col for col in result_df.columns if col not in ['Supplier ID', 'Supplier Name']]
            result_df.insert(result_df.columns.get_loc('Supplier Name') + 1, 'Total_Sum', total_sum_col)
            
            # 添加列统计功能
            self.log_message("正在添加列统计...")
            # 创建统计行
            stats_row = {'Supplier ID': '总计', 'Supplier Name': ''}
            
            # 计算各列总和（使用更明确的数据转换方式）
            for col in year_month_cols:
                # 先将'-'替换为0，然后转换为float
                col_data = result_df[col].replace('-', '0')  # 先替换为字符串'0'
                col_data = pd.to_numeric(col_data, errors='coerce')  # 转换为数值
                stats_row[col] = col_data.sum()
            
            # 计算Total_Sum列的总和
            stats_row['Total_Sum'] = result_df['Total_Sum'].sum()
            
            # 将统计行转换为DataFrame
            stats_df = pd.DataFrame([stats_row])
            
            # 将统计行插入到第二行
            result_df = pd.concat([result_df.iloc[:0], stats_df, result_df.iloc[0:]]).reset_index(drop=True)
            
            # 生成输出文件名
            latest_yearmonth = sorted_columns[0]
            output_file = f"{latest_yearmonth}_AP_Aging_Report.xlsx"
            counter = 1
            while os.path.exists(output_file):
                output_file = f"{latest_yearmonth}_AP_Aging_Report({counter}).xlsx"
                counter += 1

            # 保存文件
            self.log_message("正在保存文件...")
            with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
                result_df.to_excel(writer, index=False, sheet_name='Aggregated Data')
                
                # 样式设置
                workbook = writer.book
                worksheet = writer.sheets['Aggregated Data']
                
                # 定义右对齐样式
                right_alignment = Alignment(horizontal='right')

                # 创建会计专用格式样式
                accounting_style = NamedStyle(name="accounting_style")
                # 修改数字格式，添加负数红色显示
                accounting_style.number_format = '_ * #,##0.00_ ;[Red]_ * -#,##0.00_ ;_ * "-"??_ ;_ @_ '
                accounting_style.font = Font(name='微软雅黑', size=10)
                accounting_style.alignment = right_alignment

                # 创建对齐样式
                center_alignment = Alignment(horizontal='center', vertical='center')
                right_alignment = Alignment(horizontal='right', vertical='center')
                
                # 设置表头样式
                header_fill = PatternFill(start_color="002060", end_color="002060", fill_type="solid")
                header_font = Font(name='微软雅黑', size=9, bold=True, color='FFFFFF')
                
                # 应用表头样式
                for cell in worksheet[1]:
                    cell.fill = header_fill
                    cell.font = header_font
                    cell.alignment = right_alignment  # 添加表头右对齐

                # 设置统计行样式（第二行）
                stats_fill = PatternFill(start_color="DDEBF7", end_color="DDEBF7", fill_type="solid")
                stats_font = Font(name='微软雅黑', size=9, bold=True, color='002060')
                
                # 先应用会计格式和字体样式
                for cell in worksheet[2]:  # 第二行是统计行
                    if isinstance(cell.value, (int, float)):  # 对统计行的数值应用会计格式
                        cell.style = accounting_style
                    cell.font = stats_font
                    cell.alignment = right_alignment  # 统计行右对齐
                
                # 最后应用背景颜色
                for cell in worksheet[2]:
                    cell.fill = stats_fill

                # 创建数据行字体样式
                data_font = Font(name='微软雅黑', size=10)

                # 应用样式到所有数据行
                for row in worksheet.iter_rows(min_row=3, max_row=worksheet.max_row):
                    for cell in row:
                        cell.font = data_font  # 设置数据行字体
                        if isinstance(cell.value, (int, float)):  # 数值类型应用会计格式
                            cell.style = accounting_style
                        cell.alignment = right_alignment  # 所有数据行右对齐
                
                # 设置列宽
                worksheet.column_dimensions['A'].width = 19
                worksheet.column_dimensions['B'].width = 40
                for idx, col in enumerate(worksheet.columns, start=1):
                    if idx > 2:
                        worksheet.column_dimensions[col[0].column_letter].width = 20
                
                # 设置行高
                for row in worksheet.iter_rows(min_row=1, max_row=worksheet.max_row):
                    worksheet.row_dimensions[row[0].row].height = 22.5
                
                # 冻结表头
                worksheet.freeze_panes = worksheet['A2']
                
                # 隐藏网格线
                worksheet.sheet_view.showGridLines = False

                # 在表头上方插入新行
                worksheet.insert_rows(1)  # 在第一行插入新行

                # 设置新行的样式（使用统计行样式）
                stats_fill = PatternFill(start_color="DDEBF7", end_color="DDEBF7", fill_type="solid")
                stats_font = Font(name='微软雅黑', size=9, bold=True, color='002060')
                
                # 在A1写入生成时间
                now = datetime.now()
                worksheet['A1'].value = now.strftime('%Y-%m-%d %H:%M:%S')
                worksheet['A1'].fill = stats_fill
                worksheet['A1'].font = stats_font
                worksheet['A1'].alignment = right_alignment

                # 在B1写入标题
                worksheet['B1'].value = 'AP Aging Report by Suppliers'
                worksheet['B1'].fill = stats_fill
                worksheet['B1'].font = stats_font
                worksheet['B1'].alignment = right_alignment

                # 在C1应用统计行样式
                worksheet['C1'].fill = stats_fill
                worksheet['C1'].font = stats_font
                worksheet['C1'].alignment = right_alignment

                # 设置新建行样式
                start_col = 4  # 从第4列开始
                days = 30
                for col in range(start_col, worksheet.max_column + 1):
                    cell = worksheet.cell(row=1, column=col)
                    cell.value = f"{days}Days"
                    cell.fill = stats_fill
                    cell.font = stats_font
                    cell.alignment = right_alignment
                    days += 30

                # 冻结前3行
                worksheet.freeze_panes = 'A4'  # 冻结前三行

                # 原表头行现在在第2行
                header_row = 2
            
            self.log_message(f"处理完成！文件已保存到: {output_file}")
            messagebox.showinfo("完成", f"文件处理完成！\n保存路径: {output_file}")
            
            # 提示用户是否打开文件
            open_file = messagebox.askyesno('打开文件', '文件已保存，是否立即打开？')
            if open_file:
                os.startfile(output_file)

        except Exception as e:
            self.log_message(f"处理出错: {str(e)}")
            messagebox.showerror("错误", f"处理过程中出现错误: {str(e)}")
        finally:
            self.processing = False
            self.process_btn.config(state=tk.NORMAL)
            
if __name__ == "__main__":
    root = tk.Tk()
    app = AP_Aging_Report_App(root, root)  # 将root同时作为container和master_window传递
    root.mainloop()
