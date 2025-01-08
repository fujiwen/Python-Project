import pandas as pd
import numpy as np
from datetime import datetime
import os
import glob
import time
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment

def clean_gl_data(file_path):
    df = pd.read_excel(file_path, sheet_name='sheet1', skiprows=1)
    df.columns = df.columns.str.strip().str.lower()
    
    if 'account' in df.columns:
        df['account'] = df['account'].astype(str).str.strip()
        filtered_df = df[df['account'] == '115307']
        
        new_columns = {
            'journal date': 'Date',
            'user': 'Reference',
            'line description': 'Description',
            'base amount': 'Base Amount'
        }
        cleaned_df = filtered_df[list(new_columns)].rename(columns=new_columns)
        cleaned_df['Date'] = pd.to_datetime(cleaned_df['Date']).dt.date
        
        return cleaned_df if not cleaned_df.empty else None
    return None

def process_bank_data(file_path):
    df = pd.read_excel(file_path, engine='xlrd', skiprows=8)
    default_payee_name = "海南空港开发产业集团有限公司琼中福朋喜来登酒店分公司"
    
    def convert_date_format(date_str):
        try:
            return datetime.strptime(str(date_str), '%Y%m%d').strftime('%Y-%m-%d')
        except ValueError:
            return date_str
    
    new_rows = []
    for _, row in df.iterrows():
        transaction_date = convert_date_format(row.get('交易日期[ Transaction Date ]', ''))
        payee_name = row.get('收款人名称[ Payee\'s Name ]', np.nan)
        if pd.isna(payee_name) or str(payee_name).strip() == '':
            payee_name = default_payee_name
        
        trade_amount = float(row.get('交易金额[ Trade Amount ]', 0.0))
        debit_credit = "收款" if trade_amount > 0 else "付款" if trade_amount < 0 else ""
        
        new_row = {
            '日期': transaction_date,
            '对方户名': payee_name,
            '用途': row.get('用途[ Purpose ]', ''),
            '交易金额': trade_amount,
            '借方/贷方': debit_credit,
            '交易流水号': str(row.get('交易流水号[ Transaction reference number ]', ''))
        }
        new_rows.append(new_row)
    
    return pd.DataFrame(new_rows)

def main():
    gl_files = glob.glob('gl*.xlsx')
    bank_files = glob.glob('bank*.xls')
    combined_file_path = 'Combined_Data.xlsx'

    with pd.ExcelWriter(combined_file_path, engine='openpyxl') as writer:
        if gl_files:
            gl_data = clean_gl_data(gl_files[0])
            if gl_data is not None and not gl_data.empty:
                gl_data.to_excel(writer, sheet_name='GL Data', index=False)
        
        if bank_files:
            bank_data = process_bank_data(bank_files[0])
            if not bank_data.empty:
                bank_data.to_excel(writer, sheet_name='Bank Data', index=False)
    
    print(f"所有数据已成功合并到 {combined_file_path}")

if __name__ == "__main__":
    main()

# 添加5秒延迟
time.sleep(5)

# 第二段代码开始
file_path = 'Combined_Data.xlsx'
gl_data = pd.read_excel(file_path, sheet_name='GL Data')
bank_data = pd.read_excel(file_path, sheet_name='Bank Data')

gl_data['Base Amount'] = pd.to_numeric(gl_data['Base Amount'], errors='coerce')
bank_data['交易金额'] = pd.to_numeric(bank_data['交易金额'], errors='coerce')

matches = []
unmatched_gl = []
unmatched_bank = []

bank_data_index_matched = set()
gl_data_index_matched = set()

for index_bank, row_bank in bank_data.iterrows():
    match = gl_data[(abs(gl_data['Base Amount']) == abs(row_bank['交易金额'])) & 
                    (((row_bank['交易金额'] < 0) & (gl_data['Base Amount'] < 0)) | 
                     ((row_bank['交易金额'] > 0) & (gl_data['Base Amount'] > 0)))]
    
    if not match.empty:
        for index_gl, row_gl in match.iterrows():
            formatted_date = row_gl['Date'].strftime('%Y-%m-%d') if not pd.isna(row_gl['Date']) else ''
            
            matches.append({
                '日期': row_bank['日期'],
                '对方户名': row_bank['对方户名'],
                '用途': row_bank['用途'],
                '交易流水号': str(row_bank['交易流水号']),
                '借方/贷方': row_bank['借方/贷方'],
                '交易金额': row_bank['交易金额'],
                '与总帐核对': row_gl['Reference'],
                ' ': '',  
                'Check with Bank': str(row_bank['交易流水号']),
                'Trans Date': formatted_date,
                'Description': row_gl['Description'],
                'Base Amount': row_gl['Base Amount'],
            })
            bank_data_index_matched.add(index_bank)
            gl_data_index_matched.add(index_gl)
            break  

for index_gl, row_gl in gl_data.iterrows():
    if index_gl not in gl_data_index_matched:
        unmatched_gl.append({
            'Trans Date': row_gl['Date'].strftime('%Y-%m-%d') if not pd.isna(row_gl['Date']) else '',
            'Description': row_gl['Description'],
            'Base Amount': row_gl['Base Amount'],
            'Reference': row_gl['Reference']
        })

for index_bank, row_bank in bank_data.iterrows():
    if index_bank not in bank_data_index_matched:
        unmatched_bank.append({
            '日期': row_bank['日期'],
            '对方户名': row_bank['对方户名'],
            '用途': row_bank['用途'],
            '借方/贷方': row_bank['借方/贷方'],
            '交易金额': row_bank['交易金额'],
            '交易流水号': str(row_bank['交易流水号'])
        })

verify_data = pd.DataFrame(matches)

unmatched_gl_df = pd.DataFrame(unmatched_gl)
unmatched_bank_df = pd.DataFrame(unmatched_bank)

with pd.ExcelWriter(file_path, mode='a', engine='openpyxl', if_sheet_exists='replace') as writer:
    verify_data.to_excel(writer, sheet_name='Bank_OK', index=False)
    unmatched_bank_df.to_excel(writer, sheet_name='Unmatched_Bank_Data', index=False)
    unmatched_gl_df.to_excel(writer, sheet_name='Unmatched_GL_Data', index=False)
    

wb = load_workbook(file_path)

green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
green_font = Font(color="006100")

ws_verify = wb['Bank_OK']

ws_verify.insert_rows(1)
ws_verify.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(verify_data.columns))
ws_verify['A1'] = "银行 核对已成功 明细"
ws_verify['A1'].font = green_font
ws_verify['A1'].fill = green_fill

yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
yellow_font = Font(color="000000")  

ws_unmatched_gl = wb['Unmatched_GL_Data']
ws_unmatched_gl.insert_rows(1)
ws_unmatched_gl.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(unmatched_gl_df.columns))
ws_unmatched_gl['A1'] = "未匹配GL_DATA"
ws_unmatched_gl['A1'].font = yellow_font
ws_unmatched_gl['A1'].fill = yellow_fill

ws_unmatched_bank = wb['Unmatched_Bank_Data']
ws_unmatched_bank.insert_rows(1)
ws_unmatched_bank.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(unmatched_bank_df.columns))
ws_unmatched_bank['A1'] = "未匹配BANK_DATA"
ws_unmatched_bank['A1'].font = yellow_font
ws_unmatched_bank['A1'].fill = yellow_fill

def unmerge_header(worksheet):
    merged_ranges = worksheet.merged_cells.ranges.copy()
    for merged_range in merged_ranges:
        worksheet.unmerge_cells(str(merged_range))

unmerge_header(ws_verify)
unmerge_header(ws_unmatched_gl)
unmerge_header(ws_unmatched_bank)

def adjust_columns_and_alignment(worksheet, column_widths):
    for col_letter, width_info in column_widths.items():
        width, alignment_str = width_info
        worksheet.column_dimensions[col_letter].width = width
        
        alignment = None
        if alignment_str == 'center':
            alignment = Alignment(horizontal='center')
        elif alignment_str == 'left':
            alignment = Alignment(horizontal='left')
        elif alignment_str == 'right':
            alignment = Alignment(horizontal='right')
        
        for cell in worksheet[col_letter]:
            cell.alignment = alignment

column_widths = {
    'A': (22.92, 'center'),     
    'B': (42.26, 'left'),      
    'C': (42.26, 'left'),      
    'D': (13.46, 'center'),    
    'E': (13.46, 'center'),    
    'F': (13.46, 'right'),     
    'G': (17, 'center'),    
    'H': (1, 'center'),    
    'I': (17, 'center'),    
    'J': (22.92, 'left'),      
    'K': (42.26, 'right'),     
    'L': (13.46, 'right')      
}

adjust_columns_and_alignment(ws_verify, column_widths)
adjust_columns_and_alignment(ws_unmatched_gl, column_widths)
adjust_columns_and_alignment(ws_unmatched_bank, column_widths)

def set_header_style(worksheet, header_styles, header_row=2):
    for col_letter, style_info in header_styles.items():
        background_color, font_color, font_name, font_size = style_info
        cell = worksheet[f'{col_letter}{header_row}']
        cell.fill = PatternFill(start_color=background_color, end_color=background_color, fill_type="solid")
        cell.font = Font(color=font_color, name=font_name, size=font_size, bold=True)
        cell.alignment = Alignment(horizontal='center')

header_styles_verify = {
    'A': ('00009B', 'FFFFFF', '微软雅黑', 10),   
    'B': ('00009B', 'FFFFFF', '微软雅黑', 10),   
    'C': ('00009B', 'FFFFFF', '微软雅黑', 10),   
    'D': ('00009B', 'FFFFFF', '微软雅黑', 10),   
    'E': ('00009B', 'FFFFFF', '微软雅黑', 10),   
    'F': ('00009B', 'FFFFFF', '微软雅黑', 10),   
    'G': ('00009B', 'FFFFFF', '微软雅黑', 10),   
    'H': ('FFFFFF', 'FFFFFF', '微软雅黑', 10),   
    'I': ('333F4F', 'FFFFFF', '微软雅黑', 10),   
    'J': ('333F4F', 'FFFFFF', '微软雅黑', 10),   
    'K': ('333F4F', 'FFFFFF', '微软雅黑', 10),   
    'L': ('333F4F', 'FFFFFF', '微软雅黑', 10),   
}

set_header_style(ws_verify, header_styles_verify)

def set_data_style(worksheet, data_styles, start_row=3):
    for row in worksheet.iter_rows(min_row=start_row):
        for cell in row:
            col_letter = cell.column_letter
            if col_letter in data_styles:
                background_color, font_color, font_name, font_size = data_styles[col_letter]
                cell.fill = PatternFill(start_color=background_color, end_color=background_color, fill_type="solid")
                cell.font = Font(color=font_color, name=font_name, size=font_size)

data_styles_verify = {
    'A': ('FFFFFFFF', '002060', '微软雅黑', 10),  
    'B': ('FFFFFFFF', '002060', '微软雅黑', 10),  
    'C': ('FFFFFFFF', '002060', '微软雅黑', 10),  
    'D': ('FFFFFFFF', '002060', '微软雅黑', 10),  
    'E': ('FFFFFFFF', '002060', '微软雅黑', 10),  
    'F': ('FFFFFFFF', '002060', '微软雅黑', 10),  
    'G': ('FFFFFFFF', '002060', '微软雅黑', 10),  
    'H': ('FFFFFFFF', '002060', '微软雅黑', 10),  
    'I': ('FFFFFFFF', '000000', '微软雅黑', 10),  
    'J': ('FFFFFFFF', '000000', '微软雅黑', 10),  
    'K': ('FFFFFFFF', '000000', '微软雅黑', 10),  
    'L': ('FFFFFFFF', '000000', '微软雅黑', 10),  
}

set_data_style(ws_verify, data_styles_verify)

header_styles_unmatched_gl = {
    'A': ('333F4F', 'FFFFFF', '微软雅黑', 10),   
    'B': ('333F4F', 'FFFFFF', '微软雅黑', 10),   
    'C': ('333F4F', 'FFFFFF', '微软雅黑', 10),   
    'D': ('333F4F', 'FFFFFF', '微软雅黑', 10),   
}

set_header_style(ws_unmatched_gl, header_styles_unmatched_gl)

data_styles_unmatched_gl = {
    'A': ('FFFFFFFF', '002060', '微软雅黑', 10),  
    'B': ('FFFFFFFF', '002060', '微软雅黑', 10),  
    'C': ('FFFFFFFF', '002060', '微软雅黑', 10),  
    'D': ('FFFFFFFF', '002060', '微软雅黑', 10),  
}

set_data_style(ws_unmatched_gl, data_styles_unmatched_gl)

header_styles_unmatched_bank = {
    'A': ('333F4F', 'FFFFFF', '微软雅黑', 10),   
    'B': ('333F4F', 'FFFFFF', '微软雅黑', 10),   
    'C': ('333F4F', 'FFFFFF', '微软雅黑', 10),   
    'D': ('333F4F', 'FFFFFF', '微软雅黑', 10),   
    'E': ('333F4F', 'FFFFFF', '微软雅黑', 10),   
    'F': ('333F4F', 'FFFFFF', '微软雅黑', 10),   
}

set_header_style(ws_unmatched_bank, header_styles_unmatched_bank)

data_styles_unmatched_bank = {
    'A': ('FFFFFFFF', '002060', '微软雅黑', 10),  
    'B': ('FFFFFFFF', '002060', '微软雅黑', 10),  
    'C': ('FFFFFFFF', '002060', '微软雅黑', 10),  
    'D': ('FFFFFFFF', '002060', '微软雅黑', 10),  
    'E': ('FFFFFFFF', '002060', '微软雅黑', 10),  
    'F': ('FFFFFFFF', '002060', '微软雅黑', 10),  
}

set_data_style(ws_unmatched_bank, data_styles_unmatched_bank)

ws_verify.freeze_panes = ws_verify['A3']
ws_unmatched_gl.freeze_panes = ws_unmatched_gl['A3']
ws_unmatched_bank.freeze_panes = ws_unmatched_bank['A3']

if 'GL Data' in wb.sheetnames:
    ws_gl_data = wb['GL Data']
    ws_gl_data.sheet_state = 'hidden'

if 'Bank Data' in wb.sheetnames:
    ws_bank_data = wb['Bank Data']
    ws_bank_data.sheet_state = 'hidden'

wb.save(file_path)

print("The 'GL Data' and 'Bank Data' sheets have been hidden.")
print("Verification completed and sorted by date in descending order. A green title has been added above the headers of the new 'Verify' sheet.")
print("Unmatched GL Data and Bank Data have been written to separate sheets named 'Unmatched_GL_Data' and 'Unmatched_Bank_Data'.")
print("Yellow titles have been added at the top of each unmatched data sheet indicating '未匹配GL_DATA' or '未匹配BANK_DATA'.")
